import os
import io
from typing import Optional, List, Dict

import logging
import pandas as pd
import requests
import xml.etree.ElementTree as ET
from bs4 import BeautifulSoup
from datetime import datetime
from fastapi import Depends, HTTPException
from fastapi.responses import StreamingResponse
from io import BytesIO
from logging.handlers import TimedRotatingFileHandler

from openpyxl.styles import PatternFill, Font, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from sqlalchemy import func, or_
from sqlalchemy.orm import Session

from auth.auth_service import AuthService, oauth2_scheme
from crud.comment_service import CommentService
from crud.user_service import UserService
from db.database import get_db
from exceptions.exceptions import (
    DatabaseQueryException,
    ExternalServiceUnavailableException,
    ForbiddenException,
    InsufficientFundsException,
    InvalidDateFormatException,
    InvalidUserDataException,
    PaymentNotFoundException,
    PaymentUpdateFailedException,
    UserNotFoundException,
)
from logs.logger_service import LoggerService
from models.payment import Payment
from models.user import User
from schemas.comment import CommentCreate
from schemas.payment import PaymentUpdate
from schemas.user import RegionEnum

log_service = LoggerService("logs/payment_service")
success_logger, error_logger = log_service.configure_loggers("payment_service_success", "payment_service_error")

class PaymentService:
    def __init__(self, db: Session = Depends(get_db)):
        self.db = db
        self.SUCCESS_STATUS = "0"
        self.user_service = UserService(db)
        self.auth_service = AuthService(db)
        self.comment_service = CommentService(db)

    def get_payments(
            self,
            cursor: Optional[int] = None,
            per_page: int = 10,
            start_date: Optional[str] = None,
            end_date: Optional[str] = None,
            ls_abon: Optional[str] = None,
            name: Optional[str] = None):

        try:
            query = self.db.query(Payment)

            if start_date or end_date:
                try:
                    if start_date:
                        parsed_start_date = datetime.strptime(start_date, '%d/%m/%Y').date()
                        query = query.filter(Payment.payment_date >= parsed_start_date)
                    if end_date:
                        parsed_end_date = datetime.strptime(end_date, '%d/%m/%Y').date()
                        query = query.filter(Payment.payment_date <= parsed_end_date)
                except ValueError as e:
                    error_logger.error(f"Invalid date format: {e}")
                    raise InvalidDateFormatException("Use dd/mm/yyyy format.")

            if ls_abon:
                query = query.filter(Payment.ls_abon.ilike(f"%{ls_abon}%"))

            if name:
                query = query.join(User, Payment.user_id == User.id).filter(
                    or_(User.name.ilike(f"%{name}%"), User.surname.ilike(f"%{name}%")))

            if cursor:
                query = query.filter(Payment.id > cursor)

            # Подсчет общего количества платежей до применения лимита
            total_count = query.count()

            # Выполняем запрос с лимитом и сортировкой
            db_payments = query.order_by(Payment.id).limit(per_page + 1).all()

            # Преобразование db_payments для включения FullName
            transformed_payments = []
            for payment in db_payments[:per_page]:
                db_user = self.db.query(User).filter(User.id == payment.user_id).first()
                transformed_payment = {
                    "id": payment.id,
                    "payment_accept": payment.payment_accept,
                    "money": payment.money,
                    "user_id": payment.user_id,
                    "annulment": payment.annulment,
                    "document_number": payment.document_number,
                    "updated_at": payment.updated_at,
                    "payment_date": payment.payment_date,
                    "payment_number": payment.payment_number,
                    "ls_abon": payment.ls_abon,
                    "payment_status": payment.payment_status,
                    "comment": payment.comment,
                    "FullName": f"{db_user.name} {db_user.surname}"  # Добавить поле FullName
                }
                transformed_payments.append(transformed_payment)

            # Определяем next_cursor на основе наличия дополнительных элементов
            next_cursor = db_payments[per_page].id if len(db_payments) > per_page else None

            success_logger.info(f"Retrieved {len(transformed_payments)} payments")
            return transformed_payments, total_count, next_cursor

        except Exception as e:
            error_logger.error(f"Error retrieving payments: {str(e)}")
            raise DatabaseQueryException(str(e))

    def single_user_payments(self,
                             login: str,
                             start_date: Optional[str] = None,
                             end_date: Optional[str] = None,
                             cursor: Optional[int] = None,
                             per_page: int = 10):
        try:
            user = self.user_service.get_user_by_login(login)
            query = self.db.query(Payment).filter(Payment.user_id == user.id)

            if cursor:
                query = query.filter(Payment.id > cursor)

            # Подсчет общего количества записей до применения лимита
            total_count = query.filter(Payment.payment_status == "Выполнен").count()

            if start_date or end_date:
                try:
                    if start_date:
                        parsed_start_date = datetime.strptime(start_date, '%Y-%m-%d').strftime('%d/%m/%Y')
                        query = query.filter(
                            Payment.payment_date >= datetime.strptime(parsed_start_date, '%d/%m/%Y').date())
                    if end_date:
                        parsed_end_date = datetime.strptime(end_date, '%Y-%m-%d').strftime('%d/%m/%Y')
                        query = query.filter(
                            Payment.payment_date <= datetime.strptime(parsed_end_date, '%d/%m/%Y').date())
                except ValueError as e:
                    error_logger.error(f"Invalid date format: {e}")
                    raise InvalidDateFormatException("Use yyyy-mm-dd format.")

            # Выполняем запрос с лимитом и сортировкой
            query = query.filter(Payment.payment_status == "Выполнен").order_by(Payment.id).limit(per_page + 1)

            payments = query.all()

            # Определяем next_cursor
            next_cursor = payments[per_page].id if len(payments) > per_page else None

            # Возвращаем только `per_page` элементов
            payments = payments[:per_page]

            success_logger.info(f"Retrieved {len(payments)} payments for user {login}")
            return payments, total_count, next_cursor

        except Exception as e:
            error_logger.error(f"Error retrieving payments for user {login}: {str(e)}")
            raise DatabaseQueryException(str(e))

    def get_payment_by_id(self, payment_id: int):
        try:
            payment = self.db.query(Payment).filter(Payment.id == payment_id).first()
            if payment is None:
                error_logger.warning(f"Payment with id {payment_id} not found")
                raise PaymentNotFoundException(payment_id)

            success_logger.info(f"Retrieved payment with id {payment_id}")
            return payment
        except Exception as e:
            error_logger.error(f"Error retrieving payment: {str(e)}")
            raise DatabaseQueryException(str(e))

        # service_type, comment

    def create_payment(self, ls_abon: int, money: int, token: str):
        try:
            login = self.auth_service.get_current_user(token=token).login

            db_user = self.db.query(User).filter(User.login == login).first()

            if db_user.access_to_payments is False:
                error_logger.warning(f"User {login} attempted to create payment without access")
                raise ForbiddenException("У вас нет доступа к оплатам!")

            user_id = db_user.id
            # name = db_user.name
            # surname = db_user.surname
            available_balance = db_user.available_balance
            spent_money = db_user.spent_money
            # region = db_user.region
            # access_to_payments = db_user.access_to_payments
            # hashed_password = db_user.hashed_password

            now = datetime.now()

            service_id_hydra = now.strftime("%Y%m%d%H%M%S")
            date_payment = str(datetime.now())[:-4]
            txn_id = service_id_hydra + str(ls_abon)

            if available_balance < int(money):
                error_logger.warning(f"Insufficient funds for user {login}")
                raise InsufficientFundsException("Not enough money")

            request_url = f"http://185.39.79.45:9080/localpayskynet_osmp/main?command=pay&txn_id={txn_id}&txn_date={service_id_hydra}&account={ls_abon}&sum={float(money)}"

            try:
                response = requests.post(request_url)
                response.encoding = 'utf-8'
            except:
                error_logger.error("Hydra is unavailable")
                raise ExternalServiceUnavailableException("Hydra")

            print(response.text)

            payment_accept = str(datetime.now())[:-4]

            try:
                root = ET.fromstring(response.text)
                osmp_txn_id = root.find('osmp_txn_id').text
                comment = root.find('comment').text
                summ = root.find('sum').text
                status = root.find('result').text

                result = {
                    'osmp_txn_id': osmp_txn_id,
                    'comment': comment,
                    'sum': summ,
                    'status': status
                }
                print(result)

            except:
                root = ET.fromstring(response.text)
                osmp_txn_id = root.find('osmp_txn_id').text
                summ = root.find('sum').text
                status = root.find('result').text
                result = {
                    'osmp_txn_id': osmp_txn_id,
                    'sum': summ,
                    'status': status,
                }
                print(result)

            summ = float(summ)

            if status == self.SUCCESS_STATUS:
                write_status = "Выполнен"
                new_payment = Payment(
                    payment_number=osmp_txn_id,
                    payment_date=date_payment,
                    payment_accept=payment_accept,
                    ls_abon=ls_abon,
                    money=summ,
                    payment_status=write_status,
                    user_id=user_id,
                    annulment=False,
                    document_number=None,
                    comment=None
                )

                self.db.add(new_payment)
                self.db.commit()
                self.db.refresh(new_payment)

                new_available_balance = available_balance - summ
                new_spent_balance = spent_money - summ

                db_user_update = self.user_service.get_user_by_id(user_id)
                db_user_update.available_balance = new_available_balance
                db_user_update.spent_money = new_spent_balance
                self.db.commit()
                self.db.refresh(db_user_update)

                success_logger.info(f"Created payment for user {login}, amount: {summ}")
                return result

            error_logger.warning(f"Payment creation failed for user {login}, status: {status}")
            return {"status": "error"}
        except Exception as e:
            error_logger.error(f"Error creating payment: {str(e)}")
            raise DatabaseQueryException(str(e))

    def update_payment(self, payment_id: int, payment: PaymentUpdate):
        db_payment = self.get_payment_by_id(payment_id)

        anullment_before = db_payment.annulment
        anullment_after = payment.annulment

        old_available_balance = db_payment.user.available_balance
        old_spent_money = db_payment.user.spent_money

        add_to_available_balance = db_payment.money
        add_to_spent_money = db_payment.money

        success_logger.info(
            f"Payment details before update: annulment={anullment_before}, available_balance={old_available_balance}, spent_money={old_spent_money}")

        for field in payment.dict(exclude_unset=True):
            if hasattr(db_payment, field):
                setattr(db_payment, field, getattr(payment, field))
        success_logger.info(f"Updated fields for payment with ID {payment_id}")

        if anullment_before != anullment_after:
            db_payment.annulment = anullment_after
            db_payment.payment_status = "Аннулирован"

            text = db_payment.comment
            user_id = db_payment.user_id
            new_available_balance = old_available_balance + add_to_available_balance
            new_spent_money = old_spent_money + add_to_spent_money
            payment_type = "Аннулирован"

            comment_to_create = CommentCreate(
                user_id=user_id,
                text=text,
                payment_type=payment_type,
                old_available_balance=old_available_balance,
                old_spent_money=old_spent_money,
                add_to_available_balance=add_to_available_balance,
                add_to_spent_money=add_to_spent_money,
                new_available_balance=new_available_balance,
                new_spent_money=new_spent_money
            )

            self.comment_service.create_comment(comment_to_create)
            success_logger.info(f"Created comment for annulled payment with ID {payment_id}")

            db_user = self.user_service.get_user_by_id(user_id)
            db_user.available_balance = new_available_balance
            db_user.spent_money = new_spent_money
            success_logger.info(f"Updated user balance for user with ID {user_id}")

        try:
            self.db.commit()
            self.db.refresh(db_payment)
        except Exception as e:
            error_logger.error(f"Error updating payment {payment_id}: {str(e)}")
            self.db.rollback()
            raise PaymentUpdateFailedException(payment_id)

    def delete_payment(self, payment_id: int):
        pass  # implement after getting proper db functions

    def export_all_user_payments(self, start_date=None, end_date=None, region: RegionEnum = None):
        try:
            # Обновленный запрос для получения пользователей с фильтрацией по региону
            query = self.db.query(User)

            if region:
                query = query.filter(User.region == region.value)

            users = query.all()
            output = BytesIO()

            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book
                worksheet = workbook.add_worksheet('Платежи пользователей')
                bold = workbook.add_format({'bold': True})

                row_offset = 0
                total_sum_all_users = 0

                for user in users:
                    query = self.db.query(Payment).filter(Payment.user_id == user.id)

                    if start_date:
                        query = query.filter(Payment.payment_date >= start_date)
                    if end_date:
                        query = query.filter(Payment.payment_date <= end_date)

                    payments = query.all()

                    if not payments:
                        continue

                    user_payments = []
                    for payment in payments:
                        user_name = self.user_service.get_user_by_id(payment.user_id).name
                        user_payments.append({
                            'Номер платежа': payment.payment_number,
                            'Время проведения платежа': payment.payment_date.strftime('%Y-%m-%d %H:%M:%S'),
                            'Лицевой счет': payment.ls_abon,
                            'Сумма': payment.money,
                            'Регион': user.region,
                            'Статус платежа': payment.payment_status,
                            'Пользователь': user_name,
                            'Аннуляция': payment.annulment,
                            'Комментарий': payment.comment
                        })

                    df = pd.DataFrame(user_payments)

                    # Расчет общей суммы для пользователя
                    total_sum = sum(payment.money for payment in payments
                                    if payment.payment_status == 'Выполнен' and not payment.annulment)
                    total_sum_all_users += total_sum

                    # Запись данных в Excel
                    worksheet.write(row_offset, 0, f"Платежи пользователя: {user.name}", bold)
                    row_offset += 1

                    # Запись заголовков для каждой подтаблицы
                    for col_num, value in enumerate(df.columns.values):
                        worksheet.write(row_offset + 1, col_num, value, bold)

                    for row_num, row_data in enumerate(df.values):
                        for col_num, value in enumerate(row_data):
                            worksheet.write(row_offset + 2 + row_num, col_num, value)

                    # Запись общей суммы
                    worksheet.write(row_offset + len(df) + 3, 0, 'Общая Сумма:', bold)
                    worksheet.write(row_offset + len(df) + 3, 1, total_sum)
                    row_offset += 1

                    # Добавляем границы для таблицы
                    border_format = workbook.add_format({'border': 1})
                    worksheet.conditional_format(row_offset, 0, row_offset + len(df) + 3, len(df.columns) - 1,
                                                 {'type': 'no_blanks', 'format': border_format})

                    row_offset += len(df) + 5  # +5 для отступа между пользователями

                # Запись общей суммы для всех пользователей
                worksheet.write(row_offset + 1, 0, 'Полная общая сумма для всех пользователей:', bold)
                worksheet.write(row_offset + 1, 1, total_sum_all_users)
                if region is not None:
                    worksheet.write(row_offset + 2, 0, f'Регион: {region.value}', bold)
                date_range = f"Период: {start_date} - {end_date}" if start_date and end_date else "Период: за все время"
                worksheet.write(row_offset + 3, 0, date_range)

            output.seek(0)
            success_logger.info("Exported all user payments successfully")
            return output
        except Exception as e:
            error_logger.error(f"Error exporting all user payments: {str(e)}")
            raise DatabaseQueryException(str(e))

    def single_user_payment_report(self, user_id=None, login=None, start_date=None, end_date=None):
        try:
            if user_id is not None:
                user = self.db.query(User).filter(User.id == user_id).first()
                success_logger.info(f"Retrieved user with ID {user_id} for payment report")
            elif login is not None:
                user = self.db.query(User).filter(User.login == login).first()
                success_logger.info(f"Retrieved user with login {login} for payment report")
            else:
                error_logger.error("User ID or login must be provided")
                raise InvalidUserDataException("Необходимо указать ID пользователя или логин")

            if not user:
                error_logger.warning(f"User not found for ID {user_id} or login {login}")
                raise HTTPException(status_code=404, detail="Пользователь не найден")

            query = self.db.query(Payment).filter(Payment.user_id == user.id)

            if start_date:
                start_date_formatted = datetime.strptime(start_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                query = query.filter(Payment.payment_date >= start_date_formatted)
                success_logger.info(f"Filtered payments from start date: {start_date_formatted}")
            if end_date:
                end_date_formatted = datetime.strptime(end_date, "%d/%m/%Y").strftime("%Y-%m-%d")
                query = query.filter(Payment.payment_date <= end_date_formatted)
                success_logger.info(f"Filtered payments until end date: {end_date_formatted}")

            payments = query.all()
            success_logger.info(f"Retrieved {len(payments)} payments for user ID {user.id}")

            if not payments:
                df = pd.DataFrame({"Сообщение": ["у пользователя нет платежей"]})
                success_logger.info(f"No payments found for user ID {user.id}")
            else:
                payment_data = [{
                    'Номер платежа': payment.payment_number,
                    'Дата проведения': payment.payment_date,
                    'Лицевой счет': payment.ls_abon,
                    'Сумма': payment.money,
                    'Статус платежа': payment.payment_status,
                    'ID пользователя': payment.user_id,
                    'Аннулирование': payment.annulment,
                    'Комментарий': payment.comment,
                } for payment in payments]

                df = pd.DataFrame(payment_data)
                total_sum = df.loc[(df['Статус платежа'] == 'Выполнен') & (df['Аннулирование'] == False), 'Сумма'].sum()
                success_logger.info(f"Total sum of completed and non-annulled payments: {total_sum}")

            output = BytesIO()
            with pd.ExcelWriter(output, engine='xlsxwriter') as writer:
                workbook = writer.book
                worksheet = workbook.add_worksheet('Отчет по платежам')
                writer.sheets['Отчет по платежам'] = worksheet

                title = f"Данные по платежам: {user.name}"
                worksheet.merge_range('A1:H1', title, workbook.add_format({'align': 'center', 'bold': True}))

                df.to_excel(writer, startrow=2, index=False, sheet_name='Отчет по платежам')

                if not payments:
                    worksheet.write(3, 0, "у пользователя нет платежей", workbook.add_format({'bold': True}))
                else:
                    row_offset = len(df) + 3
                    worksheet.write(row_offset, 1, 'Общая сумма:', workbook.add_format({'bold': True}))
                    worksheet.write(row_offset, 3, total_sum, workbook.add_format({'bold': True}))

                    date_range = f"Период: {start_date} - {end_date}" if start_date and end_date else "Период: за все время"
                    worksheet.write(row_offset + 2, 0, date_range)

                success_logger.info(f"Excel report created successfully for user ID {user.id}")

            output.seek(0)
            return output
        except Exception as e:
            error_logger.error(
                f"Error generating payment report for user ID {user_id or 'N/A'}, login {login or 'N/A'}: {str(e)}")
            raise DatabaseQueryException(str(e))

    def planup_localpay_compare(self, user_login: str, start_date: Optional[str] = None,
                                end_date: Optional[str] = None):
        try:
            # Get payments from LocalPay
            localpay_payments, _, _ = self.single_user_payments(user_login, start_date, end_date, per_page=1000000000)

            localpay_payments = [p for p in localpay_payments if int(p.money) != 0]


            # Get user data and planup payments
            db_user = self.user_service.get_user_by_login(user_login)
            planup_url = f"http://planup.skynet.kg:8000/planup/localpay_naryd/"
            planup_id = db_user.planup_id
            data = {"planup_id": planup_id}
            if start_date:
                data["start_date"] = start_date
            if end_date:
                data["end_date"] = end_date

            response = requests.post(planup_url, data=data)
            if response.status_code != 200:
                error_logger.error(f"Failed to fetch data from Planup. Status code: {response.status_code}")
                error_logger.error(f"Response content: {response.text}")
                raise Exception(f"Failed to fetch data from Planup. Status code: {response.status_code}")

            planup_payments = response.json()

            print(planup_payments)

            # Convert planup payments to standard format and filter out zero payments
            planup_payments = [
                {
                    "id": p["id"],
                    "ls_abon": p["ls_abon"],
                    "money": int(p["money"]) if isinstance(p["money"], str) and p["money"].isdigit() and int(p["money"]) != 0 else 0
                }
                for p in planup_payments if isinstance(p["money"], str) and p["money"].isdigit() and int(p["money"]) != 0
            ]

            localpay_set = {(p.ls_abon, int(p.money)) for p in localpay_payments}
            planup_set = {(p["ls_abon"], p["money"]) for p in planup_payments}

            report = []
            localpay_total = 0

            for payment in localpay_payments:
                ls_abon = payment.ls_abon
                money = int(payment.money)
                localpay_total += money
                match = (ls_abon, money) in planup_set
                if match:
                    planup_payment = next(p for p in planup_payments if p["ls_abon"] == ls_abon and p["money"] == money)
                    report.append({
                        "ls_abon": ls_abon,
                        "localpay_money": money,
                        "planup_money": planup_payment["money"]
                    })
                else:
                    report.append({
                        "ls_abon": ls_abon,
                        "localpay_money": money,
                        "planup_money": "ТАКОЙ ОТСУТСВУЕТ В PLANUP"
                    })

            # Process planup payments
            for payment in planup_payments:
                ls_abon = payment["ls_abon"]
                money = payment["money"]
                if (ls_abon, money) not in localpay_set:
                    report.append({
                        "ls_abon": ls_abon,
                        "localpay_money": "ПЛАТЕЖ ОТСУТСВУЕТ В LOCALPAY",
                        "planup_money": money
                    })

            # Adding total sums
            planup_total = self._cumulative_sum(planup_payments)
            report.append({
                "ls_abon": "Итого",
                "localpay_money": localpay_total,
                "planup_money": planup_total
            })

            full_name = f"{db_user.name} {db_user.surname}"

            # Create the period text
            if start_date and end_date:
                period_text = f"Период: {start_date} - {end_date}"
            elif start_date and not end_date:
                period_text = f"Период: {start_date} - текущий день"
            elif not start_date and end_date:
                period_text = f"Период: самое начало - {end_date}"
            else:
                period_text = "Период: все время"

            # Создаем новую книгу Excel
            wb = Workbook()
            ws = wb.active
            ws.title = 'Payment Comparison'

            # Добавляем ФИО пользователя
            ws['A1'] = f"Отчет Платежей {full_name}"
            ws.merge_cells('A1:C1')
            ws['A1'].font = Font(bold=True, size=14)

            # Добавляем заголовки
            headers = ['ЛС абонента', 'LocalPay сумма', 'Planup сумма']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=3, column=col, value=header)
                cell.font = Font(bold=True)

            # Определяем стиль границ
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Заполняем данные и применяем форматирование
            red_fill = PatternFill(start_color='FFCCCB', end_color='FFCCCB', fill_type='solid')
            yellow_fill = PatternFill(start_color='FFFFCC', end_color='FFFFCC', fill_type='solid')
            bold_font = Font(bold=True)

            for row, payment in enumerate(report, start=4):
                for col in range(1, 4):
                    cell = ws.cell(row=row, column=col)
                    cell.border = thin_border

                ws.cell(row=row, column=1, value=payment['ls_abon'])
                ws.cell(row=row, column=2, value=payment['localpay_money'])
                ws.cell(row=row, column=3, value=payment['planup_money'])

                if payment['planup_money'] == "ТАКОЙ ОТСУТСВУЕТ В PLANUP":
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).fill = red_fill

                if payment['localpay_money'] == "ПЛАТЕЖ ОТСУТСВУЕТ В LOCALPAY":
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).fill = yellow_fill

                if payment['ls_abon'] == "Итого":
                    for col in range(1, 4):
                        ws.cell(row=row, column=col).font = bold_font

            # Adding the period information at the end
            last_row = ws.max_row + 2
            ws.cell(row=last_row, column=1, value=period_text).font = bold_font
            ws.merge_cells(start_row=last_row, start_column=1, end_row=last_row, end_column=3)

            # Добавляем границы к заголовкам
            for col in range(1, 4):
                ws.cell(row=3, column=col).border = thin_border

            # Автоподбор ширины столбцов
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width

            # Сохраняем в BytesIO объект
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            return output, full_name

        except Exception as e:
            error_logger.error(f"Error comparing user payments with Planup: {str(e)}")
            raise DatabaseQueryException(str(e))


    def _cumulative_sum(self, payments: List[Dict]) -> float:
        return sum(float(payment['money']) if isinstance(payment['money'], (int, float)) else 0 for payment in payments)

